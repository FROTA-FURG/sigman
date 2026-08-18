"""
Extração corrigida do planejamento da Lancha Larus (Plano52LL.xlsx).

O main.py original assume o layout de colunas do AS/CMI (B=tag, C=descrição,
D=atividade, E=periodicidade, semanas a partir de F) e usa a coluna B como
critério de "linha tem equipamento?". A planilha da LL tem um layout
diferente:

  B = rótulo de sistema (só preenchido na 1ª linha de cada grupo de tags)
  C = tag
  D = descrição do equipamento
  E = atividade
  F = periodicidade (a própria célula também é colorida com a cor da legenda)
  G = coluna em branco (espaçador)
  H em diante = semanas (numeradas 1.0, 2.0, 3.0...)

Como o main.py checava `row[0].value is None` (coluna B) pra decidir se a
linha tinha dado, e a coluna B só vem preenchida na primeira linha de cada
grupo, ~85% das linhas reais (as continuações de cada grupo) eram
descartadas silenciosamente. Este script corrige o deslocamento de colunas,
usa a coluna C como critério real de "tem tag", e começa a varredura de cor
na coluna H (não F, que é a própria célula de periodicidade).

Também trata cores por tema do Excel (algumas células de semana da LL usam
cor indexada no tema do workbook, não RGB direto) — resolvidas via o tema do
arquivo. Tema índice 9 = RGB 70AD47, muito próximo (mas não idêntico) do
verde da legenda pra "Mensal" (6AA84F) -> tratado como Mensal, mas é uma
suposição que precisa confirmação.

Uso:
    .venv/bin/python extract_ll_fixed.py
"""
import json
import re

import openpyxl

ARQUIVO_ENTRADA = "Plano52LL.xlsx"
ARQUIVO_SAIDA = "planejamentoLL.json"

# Cores por tema encontradas nas células de semana da LL -> nome da legenda.
# Resolvido a partir do tema do workbook (índice 9 = RGB 70AD47, muito
# próximo do verde 6AA84F usado pra "Mensal" na legenda RGB) -- suposição,
# reportada ao usuário pra confirmar.
THEME_COLOR_MAP = {
    9: "Mensal",
}

# Índices de tema que aparecem nas colunas de semana mas são só formatação
# de fundo (célula sempre vazia, sem "x" nem nenhuma marca real) -- 461
# ocorrências confirmadas, todas com valor=None, então tratadas como
# "sem marca" igual ao branco/00000000 do RGB.
THEME_SEM_MARCA = {0}

# Cores que aparecem nas células de semana mas não têm nenhum mapeamento
# confiável (nem legenda, nem tema conhecido) -- mantidas como o próprio
# código bruto no JSON, pra ficarem visíveis e serem decididas depois,
# igual ao "Indefinido" do AS.
SEM_MARCA = {"00000000", "FFFFFFFF"}


def resolve_theme_colors(wb):
    theme_xml = wb.loaded_theme
    if isinstance(theme_xml, bytes):
        theme_xml = theme_xml.decode("utf-8")
    return re.findall(r'<a:srgbClr val="([0-9A-Fa-f]{6})"/>', theme_xml or "")


wb = openpyxl.load_workbook(ARQUIVO_ENTRADA)
sheet = wb.active

cabecalhos = []
plano_manutencao = []
dicionario_legendas = {}

# PASSO 1: legenda (mesma lógica do main.py, cobre min_row=6..210 -- a
# legenda da LL está na linha 159, dentro desse range).
for row in sheet.iter_rows(min_row=4, max_row=4, min_col=2):
    for celula in row:
        cabecalhos.append(celula.value)

for row in sheet.iter_rows(min_row=6, min_col=2, max_row=210):
    valores_linha = [str(c.value).strip().lower() for c in row if c.value is not None]
    if "legenda" in valores_linha:
        for j in range(len(row)):
            celula_atual = row[j]
            cor = celula_atual.fill.start_color.index
            if cor and cor != "00000000":
                if j + 1 < len(row):
                    texto_legenda = row[j + 1].value
                    if texto_legenda:
                        dicionario_legendas[cor] = str(texto_legenda).strip()
        break

# PASSO 2: equipamentos -- layout corrigido pra LL.
# min_row=5 (não 6): a 1ª linha de dado real da LL é a linha 5.
for row in sheet.iter_rows(min_row=5, min_col=2, max_row=210):
    valores_linha = [str(c.value).strip().lower() for c in row if c.value is not None]
    if "legenda" in valores_linha:
        continue

    tag = row[1].value  # coluna C
    if tag is None:
        continue

    equipamento = {
        "tag": tag,
        "descricao_do_eqto": row[2].value,  # coluna D
        "atividade": row[3].value,  # coluna E
        "periodicidade": row[4].value,  # coluna F
        "datas": [],
    }

    # Semanas começam na coluna H = índice 6 (B=0, C=1, D=2, E=3, F=4, G=5, H=6)
    for i in range(6, len(row)):
        celula_data = row[i]
        cor_obj = celula_data.fill.start_color

        if cor_obj.type == "theme":
            if cor_obj.index in THEME_SEM_MARCA:
                continue
            sentido_da_cor = THEME_COLOR_MAP.get(cor_obj.index, f"tema:{cor_obj.index}")
        else:
            cor = cor_obj.index
            if not cor or cor in SEM_MARCA:
                continue
            sentido_da_cor = dicionario_legendas.get(cor, cor)

        nome_da_semana = cabecalhos[i]
        equipamento["datas"].append({nome_da_semana: sentido_da_cor})

    plano_manutencao.append(equipamento)

resultado_final = {"legendas": dicionario_legendas, "equipamentos": plano_manutencao}

with open(ARQUIVO_SAIDA, "w", encoding="utf-8") as f:
    json.dump(resultado_final, f, indent=4, ensure_ascii=False)

print(f"Sucesso! {len(plano_manutencao)} linhas de equipamento extraídas em {ARQUIVO_SAIDA}")
