"""
Extração corrigida do novo plano da Atlântico Sul (Plano 52 semanas AS.xlsx,
aba "Plano de 52 Semanas AS").

Mesma classe de bug já vista na LL: o main.py original assume o layout do
plano antigo (B=tag, C=descrição, D=atividade, E=periodicidade, semanas a
partir de F). Essa planilha nova tem outro layout:

  B = rótulo de grupo (só na 1ª linha de cada equipamento, não confiável)
  C = tag
  D = descrição do equipamento
  E = atividade
  F = periodicidade
  G = Nº de Pontos (não usado)
  H em diante = semanas (formato "DD/MM à DD/MM", igual ao plano antigo)

Diferente da LL, aqui a coluna C (tag) já vem preenchida em toda linha do
grupo (não só na primeira), então não há problema de linhas perdidas por
checar a coluna errada -- só o deslocamento de colunas mesmo.

7 linhas são só um lembrete apontando pra aba "Plano de Lubrificação"
("Lubrificação do Guindaste Munck DESCRIÇÃO" etc, sem periodicidade nem
semana marcada) -- excluídas aqui. Além dessas, outras 83 linhas (dos
mesmos 7 equipamentos: AFR01, GBB, GBE, GCE01, GMO01, GMU01, GOC01) têm só
o nome do ponto de lubrificação como atividade, com periodicidade em texto
mas nenhuma semana marcada -- é o mesmo conteúdo da aba de lubrificação,
só que solto na aba principal sem cronograma. Confirmado com o usuário
excluir também (esses 7 equipamentos ficam sem nenhuma OS por enquanto,
até a lubrificação ser importada depois). A aba de lubrificação em si fica
de fora por decisão do usuário, por enquanto.

Uso:
    .venv/bin/python extract_as_v2_fixed.py
"""
import json

import openpyxl

ARQUIVO_ENTRADA = "Plano 52 semanas AS.xlsx"
ABA = "Plano de 52 Semanas AS "
ARQUIVO_SAIDA = "planejamentoAS_v2.json"


def parse_horas(valor):
    """Horas escritas na célula da semana (0.5, 4, 16...).

    A planilha usa '-' (e às vezes deixa em branco) quando não informa a
    duração da tarefa -> retorna None, e a OS fica sem horas estimadas em
    vez de receber um zero que pareceria 'tarefa instantânea'.
    """
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip().replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return None

wb = openpyxl.load_workbook(ARQUIVO_ENTRADA)
sheet = wb[ABA]

cabecalhos = []
plano_manutencao = []
dicionario_legendas = {}
pulados_placeholder = 0

for row in sheet.iter_rows(min_row=4, max_row=4, min_col=2):
    for celula in row:
        cabecalhos.append(celula.value)

# PASSO 1: legenda (mesma lógica do main.py -- a legenda dessa planilha
# está na linha 161, dentro do range min_row=6..210).
for row in sheet.iter_rows(min_row=6, min_col=2, max_row=210):
    valores_linha = [str(c.value).strip().lower() for c in row if c.value is not None]
    if "legenda" in valores_linha:
        for j in range(len(row)):
            celula_atual = row[j]
            cor = celula_atual.fill.start_color.index
            if cor and cor != "00000000":
                if j + 1 < len(row):
                    texto_legenda = row[j + 1].value
                    if texto_legenda:
                        dicionario_legendas[cor] = str(texto_legenda).strip()
        break

# PASSO 2: equipamentos -- layout corrigido (tag na coluna C, semanas a
# partir da H).
for row in sheet.iter_rows(min_row=5, min_col=2, max_row=160):
    valores_linha = [str(c.value).strip().lower() for c in row if c.value is not None]
    if "legenda" in valores_linha:
        continue

    tag = row[1].value  # coluna C
    if tag is None:
        continue
    tag = tag.strip()

    atividade = row[3].value  # coluna E
    # "DESCRIÇÃO"/"DESCRIÇÂO" (varia o acento na planilha) -- placeholder
    # apontando pra aba de lubrificação, sem conteúdo próprio.
    if atividade and atividade.strip().upper().endswith(("DESCRIÇÃO", "DESCRIÇÂO")):
        pulados_placeholder += 1
        continue

    equipamento = {
        "tag": tag,
        "descricao_do_eqto": row[2].value,  # coluna D
        "atividade": atividade,
        "periodicidade": row[4].value,  # coluna F
        "datas": [],
    }

    # Semanas começam na coluna H = índice 6 (B=0, C=1, D=2, E=3, F=4, G=5, H=6)
    for i in range(6, len(row)):
        celula_data = row[i]
        cor = celula_data.fill.start_color.index

        if cor and cor not in ("00000000", "FFFFFFFF"):
            nome_da_semana = cabecalhos[i]
            sentido_da_cor = dicionario_legendas.get(cor, cor)
            equipamento["datas"].append({
                "semana": nome_da_semana,
                "periodicidade": sentido_da_cor,
                # O número escrito dentro da célula colorida é quanto tempo a
                # tarefa leva (Hh). Algumas células trazem '-' ou vêm vazias:
                # nesses casos a planilha não informou a duração.
                "horas": parse_horas(celula_data.value),
            })

    if not equipamento["datas"]:
        # Conteúdo de lubrificação solto na aba principal, sem cronograma
        # (mesmos 7 equipamentos das linhas placeholder acima) -- fora por
        # enquanto, junto com a aba de lubrificação.
        pulados_placeholder += 1
        continue

    plano_manutencao.append(equipamento)

resultado_final = {"legendas": dicionario_legendas, "equipamentos": plano_manutencao}

with open(ARQUIVO_SAIDA, "w", encoding="utf-8") as f:
    json.dump(resultado_final, f, indent=4, ensure_ascii=False)

print(f"Sucesso! {len(plano_manutencao)} linhas extraídas ({pulados_placeholder} placeholders excluídos) em {ARQUIVO_SAIDA}")
