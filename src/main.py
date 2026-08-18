import openpyxl
import json

# 1. Carrega a planilha
# wb = openpyxl.load_workbook('Plano52AS.xlsx')
# wb = openpyxl.load_workbook('Plano52CMI.xlsx')
# wb = openpyxl.load_workbook('Plano52LL.xlsx')
wb = openpyxl.load_workbook('Plano 52 semanas AS.xlsx')
sheet = wb.active

cabecalhos = []
plano_manutencao = []
dicionario_legendas = {}

# 2. Captura os cabeçalhos (Linha 4)
for row in sheet.iter_rows(min_row=4, max_row=4, min_col=2):
    for celula in row:
        cabecalhos.append(celula.value)

# ==========================================
# PASSO 1: PRÉ-VARREDURA (Acha a Legenda primeiro!)
# ==========================================
for row in sheet.iter_rows(min_row=6, min_col=2, max_row=210):
    valores_linha = [str(c.value).strip().lower() for c in row if c.value is not None]
    
    if 'legenda' in valores_linha:
        for j in range(len(row)):
            celula_atual = row[j]
            cor = celula_atual.fill.start_color.index
            
            if cor and cor != '00000000':
                if j + 1 < len(row):
                    proxima_celula = row[j + 1]
                    texto_legenda = proxima_celula.value
                    
                    if texto_legenda:
                        dicionario_legendas[cor] = str(texto_legenda).strip()
        # Achou a legenda e preencheu o dicionário? Pode parar essa pré-varredura
        break 

# ==========================================
# PASSO 2: VARREDURA PRINCIPAL (Equipamentos)
# ==========================================
for row in sheet.iter_rows(min_row=6, min_col=2, max_row=210):
    
    # Se a linha atual for a da legenda, pula (para não virar equipamento)
    valores_linha = [str(c.value).strip().lower() for c in row if c.value is not None]
    if 'legenda' in valores_linha:
        continue

    # Se não tem TAG, pula
    if row[0].value is None: 
        continue

    equipamento = {
        "tag": row[0].value,
        "descricao_do_eqto": row[1].value,
        "atividade": row[2].value,
        "periodicidade": row[3].value,
        "datas": []
    }

    # 4. Verifica as colunas de datas
    for i in range(4, len(row)):
        celula_data = row[i]
        cor = celula_data.fill.start_color.index
        
        if cor and cor != '00000000':
            nome_da_semana = cabecalhos[i]

            if cor == 'FFFF0000':
                sentido_da_cor = 'Indefinido'
            else:
                sentido_da_cor = dicionario_legendas.get(cor, cor)
            
            equipamento["datas"].append({
                nome_da_semana: sentido_da_cor
            })
            
    plano_manutencao.append(equipamento)

# 5. Salvar o arquivo
# arquivo = 'planejamentoAS.json'
# arquivo = 'planejamentoCMI.json'
# arquivo = 'planejamentoLL.json'
arquivo = 'planejamentoAS_v2.json'

resultado_final = {
    "legendas": dicionario_legendas,
    "equipamentos": plano_manutencao
}

with open(arquivo, 'w', encoding='utf-8') as arquivo_json:
    json.dump(resultado_final, arquivo_json, indent=4, ensure_ascii=False)

print(f"Sucesso! O arquivo foi salvo em: {arquivo}")