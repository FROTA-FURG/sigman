"""
Extração do plano da Ciências do Mar 1 (Plano52CMI.xlsx, aba "PCM CMI 2024").

Layout dessa planilha (diferente da AS/LL):
  B = tag
  C = descrição do equipamento
  D = atividade
  E = periodicidade (texto livre -- às vezes vem "350h/Trimestral", com o
      intervalo por hora de uso junto do calendário; a parte de calendário
      é extraída depois em gerar_plano_cmi_servidor.py)
  F em diante = semanas (formato "DD/MM à DD/MM"), sem coluna extra entre
      periodicidade e semanas (ao contrário da AS, que tem "Nº de Pontos"
      no meio)

Legenda na linha 101 (mesmo esquema de cor-célula + texto-célula-vizinha
usado no resto do plano). Linhas 94-98 são o rodapé de métricas da própria
planilha (Necessário Total, Disponível Equipe...), sem TAG preenchida --
saem sozinhas do filtro de "tag is None", não precisam de tratamento
especial mesmo tendo um preenchimento de cor (tema, não RGB) cobrindo a
linha inteira. Linha 103 ("ítem 9") é uma nota de rodapé solta depois da
legenda, fora do range de extração (max_row=100).

Uso:
    .venv/bin/python extract_cmi_fixed.py
"""
import json

import openpyxl

ARQUIVO_ENTRADA = "Plano52CMI.xlsx"
ABA = "PCM CMI 2024"
ARQUIVO_SAIDA = "planejamentoCMI_v2.json"


def parse_horas(valor):
    """Horas escritas na célula da semana. '-' ou vazio -> sem duração informada."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip().replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return None


wb = openpyxl.load_workbook(ARQUIVO_ENTRADA, data_only=True)
sheet = wb[ABA]

cabecalhos = []
plano_manutencao = []
pulados_sem_tag = 0

for row in sheet.iter_rows(min_row=3, max_row=3, min_col=2):
    for celula in row:
        cabecalhos.append(celula.value)

# PASSO 1: legenda (linha 101, mesmo esquema cor-célula + texto-vizinho).
dicionario_legendas = {}
for row in sheet.iter_rows(min_row=6, min_col=2, max_row=110):
    valores_linha = [str(c.value).strip().lower() for c in row if c.value is not None]
    if "legenda" in valores_linha:
        for j in range(len(row)):
            celula_atual = row[j]
            cor = celula_atual.fill.start_color.index
            if cor and cor not in ("00000000",) and isinstance(cor, str):
                if j + 1 < len(row):
                    texto_legenda = row[j + 1].value
                    if texto_legenda:
                        dicionario_legendas[cor] = str(texto_legenda).strip()
        break

# PASSO 2: equipamentos -- tag na coluna B, semanas a partir da F.
# max_row=100: para antes da legenda (101) e da nota de rodapé solta (103).
for row in sheet.iter_rows(min_row=5, min_col=2, max_row=100):
    valores_linha = [str(c.value).strip().lower() for c in row if c.value is not None]
    if "legenda" in valores_linha:
        continue

    tag = row[0].value  # coluna B
    if tag is None:
        pulados_sem_tag += 1
        continue
    tag = str(tag).strip()

    equipamento = {
        "tag": tag,
        "descricao_do_eqto": row[1].value,  # coluna C
        "atividade": row[2].value,  # coluna D
        "periodicidade": row[3].value,  # coluna E
        "datas": [],
    }

    # Semanas começam na coluna F = índice 4 (B=0, C=1, D=2, E=3, F=4)
    for i in range(4, len(row)):
        celula_data = row[i]
        cor = celula_data.fill.start_color.index

        # Só cores RGB reais da legenda contam -- ignora fills por tema
        # (ex.: o rodapé de métricas nas linhas 94-98, que também não tem
        # tag e já cairia fora mesmo sem esse cuidado extra).
        if cor and isinstance(cor, str) and cor not in ("00000000", "FFFFFFFF") and cor in dicionario_legendas:
            nome_da_semana = cabecalhos[i]
            equipamento["datas"].append({
                "semana": nome_da_semana,
                "periodicidade": dicionario_legendas[cor],
                "horas": parse_horas(celula_data.value),
            })

    plano_manutencao.append(equipamento)

resultado_final = {"legendas": dicionario_legendas, "equipamentos": plano_manutencao}

with open(ARQUIVO_SAIDA, "w", encoding="utf-8") as f:
    json.dump(resultado_final, f, indent=4, ensure_ascii=False)

print(f"Sucesso! {len(plano_manutencao)} linhas extraídas ({pulados_sem_tag} sem tag, ignoradas) em {ARQUIVO_SAIDA}")
print(f"Legendas encontradas: {dicionario_legendas}")
